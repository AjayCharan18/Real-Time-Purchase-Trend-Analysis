# dashboard/app.py
import os
import logging
import time
import threading
import posixpath
from datetime import datetime, timedelta
from flask import Flask, render_template, jsonify, request, send_file, make_response
from flask_socketio import SocketIO, emit
from flask_cors import CORS
import pandas as pd
from pyspark.sql import SparkSession
from pyspark.sql import functions as F
import io
import base64
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import matplotlib
matplotlib.use('Agg')  # Use non-interactive backend
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
# Amazon API integration removed

# Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('/app/logs/dashboard.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = 'purchase-trend-analysis-secret-key'
CORS(app)
socketio = SocketIO(app, cors_allowed_origins="*")

# Config
HDFS_NAMENODE = os.getenv("HDFS_NAMENODE", "hdfs://hadoop-namenode:8020")
HDFS_DATA_PATH = os.getenv("HDFS_DATA_PATH", "/purchase-analytics/aggregated")
UPDATE_INTERVAL = 30  # seconds - reduced for better performance

# Simple cache for dashboard data
cache = {
    'products': {'data': None, 'timestamp': None},
    'departments': {'data': None, 'timestamp': None},
    'hourly': {'data': None, 'timestamp': None}
}
CACHE_TTL = 30  # seconds

# Alert system configuration
alerts_config = {
    'low_sales': {'threshold': 500, 'enabled': True, 'message': 'Low sales alert: Total transactions below threshold'},
    'high_revenue': {'threshold': 50000, 'enabled': True, 'message': 'High revenue milestone reached!'},
    'product_spike': {'threshold': 2000, 'enabled': True, 'message': 'Product demand spike detected'},
    'department_drop': {'threshold': 0.8, 'enabled': True, 'message': 'Department performance drop detected'}
}

# Cooldown (seconds) before re-emitting same alert key
ALERT_COOLDOWNS = {
    'low_sales': 900,            # 15 minutes
    'high_revenue': 1800,        # 30 minutes
    'product_spike': 900,
    'department_drop': 1200      # 20 minutes
}
MAX_ACTIVE_ALERTS = 50

# Active alerts storage
active_alerts = []

class HDFSDataReader:
    def __init__(self, hdfs_namenode):
        self.hdfs_namenode = (hdfs_namenode or "hdfs://hadoop-namenode:8020").rstrip('/')
        data_path = (HDFS_DATA_PATH or "/purchase-analytics/aggregated").strip('/')
        self._aggregate_root = data_path or "purchase-analytics/aggregated"
        self.spark = None
        self._init_spark()

    def _init_spark(self):
        """Initialize Spark session with optimized settings"""
        try:
            self.spark = SparkSession.builder \
                .appName("PurchaseTrendDashboard") \
                .config("spark.sql.adaptive.enabled", "true") \
                .config("spark.sql.adaptive.coalescePartitions.enabled", "true") \
                .config("spark.sql.adaptive.coalescePartitions.minPartitionNum", "1") \
                .config("spark.sql.adaptive.coalescePartitions.initialPartitionNum", "5") \
                .config("spark.serializer", "org.apache.spark.serializer.KryoSerializer") \
                .config("spark.sql.warehouse.dir", "/tmp/spark-warehouse") \
                .config("spark.sql.streaming.checkpointLocation", "/tmp/checkpoints") \
                .config("spark.driver.memory", "1g") \
                .config("spark.executor.memory", "1g") \
                .config("spark.driver.maxResultSize", "512m") \
                .config("spark.hadoop.fs.defaultFS", self.hdfs_namenode) \
                .getOrCreate()
            
            self.spark.sparkContext.setLogLevel("ERROR")
            logger.info("Spark session initialized for dashboard")
        except Exception as e:
            logger.error(f"Failed to initialize Spark: {e}")
            self.spark = None

    def _full_path(self, relative_path: str) -> str:
        relative = (relative_path or "").strip('/')
        return posixpath.join(self.hdfs_namenode, self._aggregate_root, relative)

    def _read_parquet(self, relative_path: str):
        if not self.spark:
            return None

        full_path = self._full_path(relative_path)
        try:
            df = self.spark.read.parquet(full_path)
            if not df.take(1):
                logger.warning(f"No records found at {full_path}")
                return None
            return df
        except Exception as e:
            logger.warning(f"Unable to read parquet data at {full_path}: {e}")
            return None

    @staticmethod
    def _filter_latest_window(df):
        try:
            latest = df.select("window_end").orderBy(F.col("window_end").desc()).limit(1).head()
            if latest and latest.window_end:
                return df.filter(F.col("window_end") == latest.window_end)
        except Exception as exc:
            logger.warning(f"Failed to determine latest window: {exc}")
        return df

    def read_product_trends(self, limit=20):
        df = self._read_parquet("by_product")
        if df is None:
            logger.warning("Product aggregates missing at expected HDFS location")
            return []

        df_latest = self._filter_latest_window(df)
        rows = (
            df_latest
            .orderBy(F.col("transaction_count").desc(), F.col("total_revenue").desc())
            .limit(limit)
            .collect()
        )

        if not rows:
            logger.warning("No product aggregates available in latest window")
            return []

        products = []
        for row in rows:
            transaction_count = int(row.transaction_count or 0)
            reorder_rate = 0
            if transaction_count > 0 and row.reorder_count is not None:
                reorder_rate = int(round((row.reorder_count / transaction_count) * 100))

            products.append({
                "product_id": row.product_id,
                "product_name": row.product_name,
                "department": row.department,
                "transaction_count": transaction_count,
                "total_quantity": int(row.total_quantity or 0),
                "quantity": int(row.total_quantity or 0),
                "total_revenue": round(float(row.total_revenue or 0.0), 2),
                "avg_price": round(float(row.avg_price or 0.0), 2),
                "reorder_count": int(row.reorder_count or 0),
                "reorder_rate": reorder_rate,
                "unique_customers": int(row.unique_customers or 0),
                "latest_purchase": row.latest_purchase.isoformat() if getattr(row, "latest_purchase", None) else None,
                "window_start": row.window_start.isoformat() if getattr(row, "window_start", None) else None,
                "window_end": row.window_end.isoformat() if getattr(row, "window_end", None) else None,
                "processing_time": row.processing_time.isoformat() if getattr(row, "processing_time", None) else None
            })

        return products

    def read_department_trends(self):
        df = self._read_parquet("by_department")
        if df is None:
            logger.warning("Department aggregates missing at expected HDFS location")
            return []

        df_latest = self._filter_latest_window(df)
        rows = (
            df_latest
            .orderBy(F.col("total_revenue").desc())
            .collect()
        )

        if not rows:
            logger.warning("No department aggregates available in latest window")
            return []

        departments = []
        for row in rows:
            departments.append({
                "department": row.department,
                "transaction_count": int(row.transaction_count or 0),
                "total_quantity": int(row.total_quantity or 0),
                "total_revenue": round(float(row.total_revenue or 0.0), 2),
                "avg_order_value": round(float(row.avg_order_value or 0.0), 2),
                "avg_price": round(float(row.avg_order_value or 0.0), 2),
                "unique_users": int(row.unique_users or 0),
                "unique_products": int(row.unique_products or 0),
                "window_start": row.window_start.isoformat() if getattr(row, "window_start", None) else None,
                "window_end": row.window_end.isoformat() if getattr(row, "window_end", None) else None,
                "processing_time": row.processing_time.isoformat() if getattr(row, "processing_time", None) else None
            })

        return departments

    def read_hourly_trends(self):
        df = self._read_parquet("by_hour")
        if df is None:
            logger.warning("Hourly aggregates missing at expected HDFS location")
            return []

        agg_df = (
            df
            .groupBy("hour_of_day")
            .agg(
                F.sum("transaction_count").alias("transaction_count"),
                F.sum("total_revenue").alias("total_revenue"),
                F.avg("avg_transaction_value").alias("avg_transaction_value"),
                F.avg("avg_transaction_value").alias("avg_price"),
                F.avg("active_users").alias("active_users"),
                F.avg("active_products").alias("active_products")
            )
            .orderBy("hour_of_day")
        )

        rows = agg_df.collect()

        if not rows:
            logger.warning("No hourly aggregates available in latest window")
            return []

        hourly = []
        for row in rows:
            hourly.append({
                "hour": int(row.hour_of_day or 0),
                "hour_of_day": int(row.hour_of_day or 0),
                "transaction_count": int(row.transaction_count or 0),
                "total_revenue": round(float(row.total_revenue or 0.0), 2),
                "avg_transaction_value": round(float(row.avg_transaction_value or 0.0), 2),
                "avg_price": round(float(row.avg_price or 0.0), 2),
                "active_users": int(round(row.active_users or 0)),
                "active_products": int(round(row.active_products or 0))
            })

        return hourly

data_reader = HDFSDataReader(HDFS_NAMENODE)

def _is_alert_in_cooldown(alert_key: str, now: datetime) -> bool:
    cooldown_seconds = ALERT_COOLDOWNS.get(alert_key, 600)
    cutoff = now - timedelta(seconds=cooldown_seconds)
    for alert in active_alerts:
        if alert.get('key') == alert_key:
            try:
                ts = datetime.fromisoformat(alert['timestamp'])
                if ts >= cutoff:
                    return True
            except Exception:
                continue
    return False

def check_alerts():
    """Check for alert conditions and generate alerts"""
    global active_alerts

    try:
        # Get current stats
        products = data_reader.read_product_trends(50)
        depts = data_reader.read_department_trends()
        
        # Calculate current metrics
        total_transactions = sum(p.get("transaction_count", 0) for p in products)
        total_revenue = sum(p.get("total_revenue", 0.0) for p in products)
        top_product = max(products, key=lambda x: x.get("transaction_count", 0)) if products else None
        
        new_alerts = []
        current_time = datetime.now()
        
        # Check low sales alert
        if alerts_config['low_sales']['enabled'] and total_transactions < alerts_config['low_sales']['threshold']:
            alert_key = 'low_sales'
            if not _is_alert_in_cooldown(alert_key, current_time):
                new_alerts.append({
                    'key': alert_key,
                    'id': f"low_sales_{int(current_time.timestamp())}",
                    'type': 'warning',
                    'title': 'Low Sales Alert',
                    'message': f"Total transactions ({total_transactions:,}) below threshold ({alerts_config['low_sales']['threshold']:,})",
                    'timestamp': current_time.isoformat(),
                    'severity': 'medium'
                })

        # Check high revenue milestone
        if alerts_config['high_revenue']['enabled'] and total_revenue > alerts_config['high_revenue']['threshold']:
            alert_key = 'high_revenue'
            if not _is_alert_in_cooldown(alert_key, current_time):
                new_alerts.append({
                    'key': alert_key,
                    'id': f"high_revenue_{int(current_time.timestamp())}",
                    'type': 'success',
                    'title': 'Revenue Milestone!',
                    'message': f"Congratulations! Revenue reached ${total_revenue:,.2f}",
                    'timestamp': current_time.isoformat(),
                    'severity': 'high'
                })

        # Check product spike
        if top_product and alerts_config['product_spike']['enabled'] and top_product.get("transaction_count", 0) > alerts_config['product_spike']['threshold']:
            alert_key = f"product_spike_{top_product['product_id']}"
            if not _is_alert_in_cooldown(alert_key, current_time):
                new_alerts.append({
                    'key': alert_key,
                    'id': f"product_spike_{int(current_time.timestamp())}",
                    'type': 'info',
                    'title': 'Product Demand Spike',
                    'message': f"{top_product['product_name']} has {top_product['transaction_count']:,} transactions - consider increasing stock!",
                    'timestamp': current_time.isoformat(),
                    'severity': 'high'
                })

        # Check department performance drop
        if len(depts) > 1:
            avg_dept_revenue = sum(d.get("total_revenue", 0) for d in depts) / len(depts)
            for dept in depts:
                dept_revenue = dept.get("total_revenue", 0)
                if dept_revenue < (avg_dept_revenue * alerts_config['department_drop']['threshold']):
                    alert_key = f"department_drop_{dept['department']}"
                    if not _is_alert_in_cooldown(alert_key, current_time):
                        new_alerts.append({
                            'key': alert_key,
                            'id': f"dept_drop_{dept['department']}_{int(current_time.timestamp())}",
                            'type': 'warning',
                            'title': 'Department Performance Drop',
                            'message': f"{dept['department'].title()} department revenue (${dept_revenue:,.2f}) is below average",
                            'timestamp': current_time.isoformat(),
                            'severity': 'medium'
                        })

        # Add new alerts and keep only recent ones (last 24 hours)
        cutoff_time = current_time - timedelta(hours=24)
        active_alerts = [alert for alert in active_alerts 
                        if datetime.fromisoformat(alert['timestamp']) > cutoff_time]

        # Add new unique alerts
        existing_ids = {alert['id'] for alert in active_alerts}
        for alert in new_alerts:
            if alert['id'] not in existing_ids:
                active_alerts.append(alert)

        # Keep most recent alerts capped to avoid UI overload
        active_alerts = sorted(active_alerts, key=lambda a: a.get('timestamp', ''), reverse=True)[:MAX_ACTIVE_ALERTS]

        logger.info(f"Alert check completed. {len(new_alerts)} new alerts, {len(active_alerts)} total active")
        
    except Exception as e:
        logger.error(f"Error checking alerts: {e}")

def generate_chart_image(chart_type, data, title):
    """Generate chart images for reports"""
    plt.style.use('default')
    fig, ax = plt.subplots(figsize=(10, 6))
    
    if chart_type == 'products':
        # Top products bar chart
        products = sorted(data, key=lambda x: x.get('transaction_count', 0), reverse=True)[:10]
        names = [p['product_name'][:20] + '...' if len(p['product_name']) > 20 else p['product_name'] for p in products]
        values = [p.get('transaction_count', 0) for p in products]
        
        bars = ax.bar(range(len(names)), values, color='#2563eb', alpha=0.8)
        ax.set_xlabel('Products')
        ax.set_ylabel('Transaction Count')
        ax.set_title(title, fontsize=14, fontweight='bold')
        ax.set_xticks(range(len(names)))
        ax.set_xticklabels(names, rotation=45, ha='right')
        
        # Add value labels on bars
        for bar, value in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(values)*0.01,
                   f'{value:,}', ha='center', va='bottom', fontsize=9)
    
    elif chart_type == 'departments':
        # Department pie chart
        names = [d['department'].title() for d in data]
        values = [d.get('total_revenue', 0) for d in data]
        colors_list = ['#2563eb', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899']
        
        wedges, texts, autotexts = ax.pie(values, labels=names, autopct='%1.1f%%', 
                                         colors=colors_list[:len(names)], startangle=90)
        ax.set_title(title, fontsize=14, fontweight='bold')
        
    elif chart_type == 'hourly':
        # Hourly trends line chart
        hours = [d.get('hour', 0) for d in data]
        transactions = [d.get('transaction_count', 0) for d in data]
        
        ax.plot(hours, transactions, marker='o', linewidth=2, markersize=6, color='#2563eb')
        ax.set_xlabel('Hour of Day')
        ax.set_ylabel('Transaction Count')
        ax.set_title(title, fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3)
        ax.set_xticks(range(0, 24, 2))
    
    plt.tight_layout()
    
    # Save to bytes
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
    img_buffer.seek(0)
    plt.close()
    
    return img_buffer

def create_pdf_report():
    """Generate comprehensive PDF report"""
    try:
        # Get data
        products = data_reader.read_product_trends(20)
        departments = data_reader.read_department_trends()
        hourly = data_reader.read_hourly_trends()
        
        # Calculate stats
        total_transactions = sum(p.get("transaction_count", 0) for p in products)
        total_revenue = sum(p.get("total_revenue", 0.0) for p in products)
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, spaceAfter=30, textColor=colors.HexColor('#2563eb'))
        heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=16, spaceAfter=12, textColor=colors.HexColor('#1e293b'))
        
        # Build story
        story = []
        
        # Title
        story.append(Paragraph("Purchase Trend Analysis Report", title_style))
        story.append(Paragraph(f"Generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("Executive Summary", heading_style))
        summary_data = [
            ['Metric', 'Value'],
            ['Total Transactions', f"{total_transactions:,}"],
            ['Total Revenue', f"${total_revenue:,.2f}"],
            ['Active Products', f"{len(products)}"],
            ['Active Departments', f"{len(departments)}"],
            ['Average Transaction Value', f"${total_revenue/total_transactions:.2f}" if total_transactions > 0 else "$0.00"]
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Top Products Chart
        story.append(Paragraph("Top Products by Transaction Count", heading_style))
        products_chart = generate_chart_image('products', products, 'Top 10 Products by Transactions')
        story.append(Image(products_chart, width=6*inch, height=3.6*inch))
        story.append(Spacer(1, 20))
        
        # Top Products Table
        story.append(Paragraph("Top 10 Products Details", heading_style))
        products_data = [['Rank', 'Product Name', 'Department', 'Transactions', 'Revenue', 'Avg Price']]
        for i, product in enumerate(products[:10], 1):
            products_data.append([
                str(i),
                product['product_name'][:30] + '...' if len(product['product_name']) > 30 else product['product_name'],
                product['department'].title(),
                f"{product.get('transaction_count', 0):,}",
                f"${product.get('total_revenue', 0):,.2f}",
                f"${product.get('avg_price', 0):.2f}"
            ])
        
        products_table = Table(products_data, colWidths=[0.5*inch, 2.5*inch, 1.2*inch, 1*inch, 1*inch, 0.8*inch])
        products_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(products_table)
        story.append(Spacer(1, 20))
        
        # Department Performance Chart
        story.append(Paragraph("Department Revenue Distribution", heading_style))
        dept_chart = generate_chart_image('departments', departments, 'Revenue by Department')
        story.append(Image(dept_chart, width=6*inch, height=3.6*inch))
        story.append(Spacer(1, 20))
        
        # Hourly Trends Chart
        story.append(Paragraph("Hourly Transaction Trends", heading_style))
        hourly_chart = generate_chart_image('hourly', hourly, 'Transactions by Hour of Day')
        story.append(Image(hourly_chart, width=6*inch, height=3.6*inch))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        logger.info("PDF report generated successfully")
        return buffer
        
    except Exception as e:
        logger.error(f"Error generating PDF report: {e}")
        return None

def create_excel_export():
    """Generate Excel export with multiple sheets"""
    try:
        # Get data
        products = data_reader.read_product_trends(50)
        departments = data_reader.read_department_trends()
        hourly = data_reader.read_hourly_trends()
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Products sheet
        ws_products = wb.create_sheet("Top Products")
        headers = ['Rank', 'Product Name', 'Department', 'Transactions', 'Quantity', 'Revenue', 'Avg Price', 'Reorder Rate']
        ws_products.append(headers)
        
        for i, product in enumerate(products, 1):
            ws_products.append([
                i,
                product['product_name'],
                product['department'],
                product.get('transaction_count', 0),
                product.get('quantity', 0),
                product.get('total_revenue', 0),
                product.get('avg_price', 0),
                f"{product.get('reorder_rate', 0)}%"
            ])
        
        # Style products sheet
        for cell in ws_products[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Departments sheet
        ws_departments = wb.create_sheet("Departments")
        dept_headers = ['Department', 'Transaction Count', 'Total Revenue', 'Average Price']
        ws_departments.append(dept_headers)
        
        for dept in departments:
            ws_departments.append([
                dept['department'].title(),
                dept.get('transaction_count', 0),
                dept.get('total_revenue', 0),
                dept.get('avg_price', 0)
            ])
        
        # Style departments sheet
        for cell in ws_departments[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Hourly trends sheet
        ws_hourly = wb.create_sheet("Hourly Trends")
        hourly_headers = ['Hour', 'Transaction Count', 'Total Revenue', 'Average Price']
        ws_hourly.append(hourly_headers)
        
        for hour_data in hourly:
            ws_hourly.append([
                f"{hour_data.get('hour', 0)}:00",
                hour_data.get('transaction_count', 0),
                hour_data.get('total_revenue', 0),
                hour_data.get('avg_price', 0)
            ])
        
        # Style hourly sheet
        for cell in ws_hourly[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="F59E0B", end_color="F59E0B", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for ws in [ws_products, ws_departments, ws_hourly]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info("Excel export generated successfully")
        return buffer
        
    except Exception as e:
        logger.error(f"Error generating Excel export: {e}")
        return None

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/products/top')
def api_top_products():
    limit = int(request.args.get('limit', 20))
    
    # Check cache first
    now = time.time()
    if (cache['products']['data'] is not None and 
        cache['products']['timestamp'] is not None and 
        now - cache['products']['timestamp'] < CACHE_TTL):
        return jsonify(cache['products']['data'][:limit])
    
    # Fetch fresh data
    data = data_reader.read_product_trends(limit)
    cache['products']['data'] = data
    cache['products']['timestamp'] = now
    return jsonify(data)

@app.route('/api/departments')
def api_departments():
    # Check cache first
    now = time.time()
    if (cache['departments']['data'] is not None and 
        cache['departments']['timestamp'] is not None and 
        now - cache['departments']['timestamp'] < CACHE_TTL):
        return jsonify(cache['departments']['data'])
    
    # Fetch fresh data
    data = data_reader.read_department_trends()
    cache['departments']['data'] = data
    cache['departments']['timestamp'] = now
    return jsonify(data)

@app.route('/api/hourly')
def api_hourly():
    # Check cache first
    now = time.time()
    if (cache['hourly']['data'] is not None and 
        cache['hourly']['timestamp'] is not None and 
        now - cache['hourly']['timestamp'] < CACHE_TTL):
        return jsonify(cache['hourly']['data'])
    
    # Fetch fresh data
    data = data_reader.read_hourly_trends()
    cache['hourly']['data'] = data
    cache['hourly']['timestamp'] = now
    return jsonify(data)

@app.route('/api/opportunities')
def api_opportunities():
    """Detect revenue opportunities and sudden trend changes"""
    try:
        products = data_reader.read_product_trends(50)
        
        # Identify revenue opportunities
        opportunities = []
        
        # High-performing products (potential for stock increase)
        top_performers = [p for p in products if p.get("transaction_count", 0) > 1200]
        for product in top_performers[:5]:
            opportunities.append({
                "type": "High Demand",
                "product": product["product_name"],
                "department": product["department"],
                "transactions": product["transaction_count"],
                "revenue": product["total_revenue"],
                "recommendation": f"Increase stock - High demand with {product['transaction_count']} transactions",
                "priority": "High"
            })
        
        # High-value, low-volume products (potential for promotion)
        high_value_low_volume = [p for p in products 
                               if p.get("avg_price", 0) > 25 and p.get("transaction_count", 0) < 800]
        for product in high_value_low_volume[:3]:
            opportunities.append({
                "type": "Promotion Opportunity",
                "product": product["product_name"],
                "department": product["department"],
                "transactions": product["transaction_count"],
                "avg_price": product["avg_price"],
                "recommendation": f"Consider promotion - High value (₹{product['avg_price']:.2f}) but low volume",
                "priority": "Medium"
            })
        
        # High reorder rate products (customer favorites)
        loyal_products = [p for p in products if p.get("reorder_rate", 0) > 50]
        for product in loyal_products[:2]:
            opportunities.append({
                "type": "Customer Favorite",
                "product": product["product_name"],
                "department": product["department"],
                "reorder_rate": f"{product['reorder_rate']}%",
                "recommendation": f"Maintain stock - High customer loyalty ({product['reorder_rate']}% reorder rate)",
                "priority": "High"
            })
        
        return jsonify({
            "opportunities": opportunities,
            "total_opportunities": len(opportunities),
            "timestamp": datetime.now().isoformat()
        })
        
    except Exception as e:
        logger.error(f"Error detecting opportunities: {e}")
        return jsonify({"opportunities": [], "total_opportunities": 0})

@app.route('/api/stats')
def api_stats():
    try:
        products = data_reader.read_product_trends(50)
        depts = data_reader.read_department_trends()
        
        logger.info(f"Stats calculation: {len(products)} products, {len(depts)} departments")
        
        # Calculate totals with error handling
        total_transactions = 0
        total_revenue = 0.0
        
        for p in products:
            total_transactions += p.get("transaction_count", 0)
            total_revenue += p.get("total_revenue", 0.0)
        
        # Calculate trend indicators for business insights
        trending_products = len([p for p in products if p.get("transaction_count", 0) > 1000])
        high_value_products = len([p for p in products if p.get("avg_price", 0) > 25])
        
        stats = {
            "total_transactions": total_transactions,
            "total_revenue": round(total_revenue, 2),
            "active_products": len(products),
            "active_departments": len(depts),
            "trending_products": trending_products,
            "high_value_products": high_value_products,
            "avg_transaction_value": round(total_revenue / total_transactions, 2) if total_transactions > 0 else 0,
            "timestamp": datetime.now().isoformat()
        }
        
        logger.info(f"Calculated stats: {stats}")
        return jsonify(stats)
        
    except Exception as e:
        logger.error(f"Error calculating stats: {e}")
        return jsonify({
            "total_transactions": 0,
            "total_revenue": 0.0,
            "active_products": 0,
            "active_departments": 0,
            "trending_products": 0,
            "high_value_products": 0,
            "avg_transaction_value": 0,
            "timestamp": datetime.now().isoformat()
        })

@app.route('/api/alerts')
def api_alerts():
    """Get current active alerts"""
    try:
        # Check for new alerts
        check_alerts()
        
        # Return active alerts sorted by severity and timestamp
        sorted_alerts = sorted(active_alerts, 
                             key=lambda x: (x['severity'] == 'high', x['timestamp']), 
                             reverse=True)
        
        return jsonify({
            "alerts": sorted_alerts,
            "total_alerts": len(active_alerts),
            "timestamp": datetime.now().isoformat()
        })
    except Exception as e:
        logger.error(f"Error fetching alerts: {e}")
        return jsonify({"alerts": [], "total_alerts": 0})

@app.route('/api/alerts/config', methods=['GET', 'POST'])
def api_alerts_config():
    """Get or update alert configuration"""
    global alerts_config
    
    if request.method == 'GET':
        return jsonify(alerts_config)
    
    elif request.method == 'POST':
        try:
            new_config = request.get_json()
            if new_config:
                alerts_config.update(new_config)
                logger.info("Alert configuration updated")
                return jsonify({"status": "success", "config": alerts_config})
            else:
                return jsonify({"status": "error", "message": "Invalid configuration"}), 400
        except Exception as e:
            logger.error(f"Error updating alert config: {e}")
            return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/alerts/<alert_id>/dismiss', methods=['POST'])
def api_dismiss_alert(alert_id):
    """Dismiss a specific alert"""
    global active_alerts
    
    try:
        active_alerts = [alert for alert in active_alerts if alert['id'] != alert_id]
        logger.info(f"Alert {alert_id} dismissed")
        return jsonify({"status": "success", "message": "Alert dismissed"})
    except Exception as e:
        logger.error(f"Error dismissing alert: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/api/export/pdf')
def export_pdf():
    """Generate and download PDF report"""
    try:
        pdf_buffer = create_pdf_report()
        if pdf_buffer:
            return send_file(
                pdf_buffer,
                as_attachment=True,
                download_name=f'purchase_trend_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
                mimetype='application/pdf'
            )
        else:
            return jsonify({"error": "Failed to generate PDF report"}), 500
    except Exception as e:
        logger.error(f"Error in PDF export endpoint: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/export/excel')
def export_excel():
    """Generate and download Excel export"""
    try:
        excel_buffer = create_excel_export()
        if excel_buffer:
            return send_file(
                excel_buffer,
                as_attachment=True,
                download_name=f'purchase_trend_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        else:
            return jsonify({"error": "Failed to generate Excel export"}), 500
    except Exception as e:
        logger.error(f"Error in Excel export endpoint: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/export/status')
def export_status():
    """Get export capabilities status"""
    return jsonify({
        "pdf_available": True,
        "excel_available": True,
        "last_generated": datetime.now().isoformat(),
        "formats": ["pdf", "excel"]
    })

# Financial API routes removed - focusing on core purchase trend analysis

def background_updates():
    while True:
        products = data_reader.read_product_trends(20)
        depts = data_reader.read_department_trends()
        hourly = data_reader.read_hourly_trends()
        
        # Check for alerts
        check_alerts()
        
        # Emit updates including alerts
        socketio.emit('product_update', {'data': products})
        socketio.emit('department_update', {'data': depts})
        socketio.emit('hourly_update', {'data': hourly})
        socketio.emit('alerts_update', {'alerts': active_alerts, 'total': len(active_alerts)})
        
        time.sleep(UPDATE_INTERVAL)

@socketio.on('connect')
def on_connect():
    emit('connection_response', {'status': 'connected'})

@socketio.on('disconnect')
def on_disconnect():
    logger.info("Client disconnected")

@socketio.on('request_update')
def on_request_update():
    products = data_reader.read_product_trends(20)
    depts = data_reader.read_department_trends()
    emit('product_update', {'data': products})
    emit('department_update', {'data': depts})

if __name__ == '__main__':
    threading.Thread(target=background_updates, daemon=True).start()
    socketio.run(app, host='0.0.0.0', port=5000, debug=True, use_reloader=False)
